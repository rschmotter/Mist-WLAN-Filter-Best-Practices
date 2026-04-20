#!/usr/bin/env python3
"""
Juniper Mist WLAN Best Practices Automation Script
====================================================
Audits all enabled Org-level WLANs against Juniper Mist best practices,
reports compliance status, and optionally remediates non-compliant WLANs.

Best Practices Checked:
  1. ARP Filtering              (arp_filter = True)
  2. Multicast/Broadcast Filter (limit_bcast = True)
  3. IPv6 NDP Disabled          (allow_ipv6_ndp = False; PASS when field absent)
  4. 802.11r Fast Transition    (auth.disable_ft = False, enterprise SSIDs only)
  5. No Duplicate SSID Names    (checked across all org WLANs)

Author : Generated for Mist WLAN Best Practices Project
Version: 2.0.0
"""

import sys
import io
import time
import json
import getpass
import datetime
import argparse
import os
import re
from collections import defaultdict

# ---------------------------------------------------------------------------
# Windows UTF-8 fix - prevents UnicodeEncodeError on cp1252 consoles
# ---------------------------------------------------------------------------
try:
    if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, io.UnsupportedOperation):
    pass

# ---------------------------------------------------------------------------
# Optional dependencies
# ---------------------------------------------------------------------------
try:
    import requests
except ImportError:
    sys.exit("ERROR: 'requests' library not found.  Run: pip install requests")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("WARNING: 'openpyxl' not found - Excel export disabled.  Run: pip install openpyxl")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SCRIPT_VERSION   = "2.0.0"
API_RATE_LIMIT   = 2000
API_CALL_WINDOW  = 3600
RATE_LIMIT_BUFFER = 50

MIST_CLOUDS = {
    "1": {"name": "Global 01 (US)",   "base": "https://api.mist.com"},
    "2": {"name": "Global 02",        "base": "https://api.gc1.mist.com"},
    "3": {"name": "Europe (EU)",      "base": "https://api.eu.mist.com"},
    "4": {"name": "APAC (AC2)",       "base": "https://api.ac2.mist.com"},
    "5": {"name": "APAC (AC5)",       "base": "https://api.ac5.mist.com"},
    "6": {"name": "Canada (CA)",      "base": "https://api.ca.mist.com"},
}

ENTERPRISE_AUTH_TYPES = {"eap", "eap-reauth", "dot1x_eap", "dot1x_cert", "dot1x"}

SLE_THRESHOLD_PCT = 90.0   # sites below this % are flagged


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
LOG_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(LOG_DIR, exist_ok=True)

_run_ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_FILE  = os.path.join(LOG_DIR, f"mist_bp_{_run_ts}.log")
DEBUG_LOG = os.path.join(LOG_DIR, f"mist_bp_debug_{_run_ts}.log")
_t0       = time.time()

def _ts():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _elapsed():
    s = int(time.time() - _t0)
    return f"{s//60:02d}m{s%60:02d}s"

def log(msg, level="INFO", console=True):
    line = f"[{_ts()}] [{level}] {msg}"
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")
    if console:
        color = {
            "INFO":  "\033[0m",
            "OK":    "\033[92m",
            "WARN":  "\033[93m",
            "ERROR": "\033[91m",
            "DEBUG": "\033[94m",
            "HEAD":  "\033[1m",
        }.get(level, "\033[0m")
        try:
            print(f"{color}{line}\033[0m")
        except UnicodeEncodeError:
            safe = line.encode(sys.stdout.encoding or "ascii", errors="replace"
                               ).decode(sys.stdout.encoding or "ascii")
            print(f"{color}{safe}\033[0m")

def debug(msg):
    with open(DEBUG_LOG, "a", encoding="utf-8") as f:
        f.write(f"[{_ts()}] [DEBUG] {msg}\n")

def section(title):
    bar = "=" * 70
    log(f"\n{bar}\n  {title}\n{bar}", level="HEAD")

def progress(current, total, label=""):
    if total == 0:
        return
    pct = int(current / total * 100)
    bar = ("#" * (pct // 5)).ljust(20)
    print(f"\r  [{bar}] {pct:3d}%  {current}/{total}  {label:<30}", end="", flush=True)
    if current == total:
        print()


# ---------------------------------------------------------------------------
# MistAPI client
# ---------------------------------------------------------------------------
class MistAPI:
    def __init__(self, base_url: str, token: str):
        self.base    = base_url.rstrip("/")
        self.token   = token
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Token {token}",
            "Content-Type":  "application/json",
        })
        self._call_count   = 0
        self._window_start = time.time()

    def _check_rate_limit(self):
        now     = time.time()
        elapsed = now - self._window_start
        if elapsed >= API_CALL_WINDOW:
            self._call_count   = 0
            self._window_start = now
        if self._call_count >= (API_RATE_LIMIT - RATE_LIMIT_BUFFER):
            wait = API_CALL_WINDOW - elapsed + 5
            log(f"Rate-limit guard: {self._call_count} calls made. Sleeping {wait:.0f}s ...", "WARN")
            time.sleep(wait)
            self._call_count   = 0
            self._window_start = time.time()

    def _request(self, method, path, payload=None, params=None, silent_404=False):
        self._check_rate_limit()
        url = f"{self.base}/api/v1{path}"
        debug(f"{method} {url}  params={params}")
        try:
            r = self.session.request(method, url, json=payload, params=params, timeout=30)
            self._call_count += 1
            debug(f"  -> {r.status_code}  calls={self._call_count}")
            if r.status_code == 404 and silent_404:
                debug(f"  404 (silent) {path}")
                return None
            if r.status_code == 429:
                retry_after = int(r.headers.get("Retry-After", 60))
                log(f"429 Too Many Requests - waiting {retry_after}s ...", "WARN")
                time.sleep(retry_after)
                return self._request(method, path, payload, params, silent_404)
            r.raise_for_status()
            return r.json() if r.text else {}
        except requests.exceptions.RequestException as e:
            log(f"API ERROR: {method} {path} -> {e}", "ERROR")
            debug(f"  Exception: {e}")
            return None

    def get(self, path, params=None, silent_404=False):
        return self._request("GET", path, params=params, silent_404=silent_404)

    def put(self, path, payload):
        return self._request("PUT", path, payload=payload)

    def delete(self, path):
        return self._request("DELETE", path)

    def get_all(self, path, params=None):
        """Generic paginated GET - returns combined list across all pages."""
        results  = []
        page     = 1
        per_page = 100
        p        = dict(params or {})
        while True:
            p["page"]  = page
            p["limit"] = per_page
            data = self.get(path, params=p)
            if data is None:
                break
            chunk = data if isinstance(data, list) else []
            results.extend(chunk)
            if len(chunk) < per_page:
                break
            page += 1
        return results

    @property
    def call_count(self):
        return self._call_count


# ---------------------------------------------------------------------------
# Interactive helpers
# ---------------------------------------------------------------------------
def ask_yn(prompt_text: str) -> bool:
    while True:
        r = input(f"\n  {prompt_text} [y/N]: ").strip().lower()
        if r in ("y", "yes"):
            return True
        if r in ("n", "no", ""):
            return False
        print("  Please enter y or n.")

def prompt_cloud() -> str:
    section("Mist Cloud Selection")
    for k, v in MIST_CLOUDS.items():
        print(f"  {k}) {v['name']}  ({v['base']})")
    while True:
        choice = input("\n  Enter cloud number [1-6]: ").strip()
        if choice in MIST_CLOUDS:
            base = MIST_CLOUDS[choice]["base"]
            log(f"Selected cloud: {MIST_CLOUDS[choice]['name']}  ->  {base}")
            return base
        print("  Invalid choice - try again.")

def prompt_org_id() -> str:
    section("Org ID")
    while True:
        org = input("  Paste your Mist Org ID (UUID): ").strip()
        if re.match(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
                    org, re.I):
            log(f"Org ID: {org}")
            return org
        print("  Does not look like a valid UUID - try again.")

def prompt_token(label="SuperUser API Token") -> str:
    print(f"\n  {label} (input hidden):")
    token = getpass.getpass("  Token: ").strip()
    if not token:
        sys.exit("ERROR: Token cannot be empty.")
    return token

def verify_auth(api: MistAPI) -> dict:
    data = api.get("/self")
    if not data:
        sys.exit("ERROR: Authentication failed - check token and cloud selection.")
    log(f"Authenticated as: {data.get('email', 'unknown')}  "
        f"(privileges: {len(data.get('privileges', []))})", "OK")
    return data


# ---------------------------------------------------------------------------
# Data collection
# ---------------------------------------------------------------------------
def get_sites(api: MistAPI, org_id: str) -> list:
    log("Fetching sites ...")
    sites = api.get_all(f"/orgs/{org_id}/sites")
    log(f"  Found {len(sites)} site(s).", "OK")
    return sites

def get_org_wlans(api: MistAPI, org_id: str) -> list:
    """Fetch all Org-level WLANs (paginated)."""
    log("Fetching org-level WLANs ...")
    wlans = api.get_all(f"/orgs/{org_id}/wlans")
    log(f"  Found {len(wlans)} org WLAN(s).", "OK")
    return wlans

def get_site_client_count(api: MistAPI, site_id: str) -> int:
    clients = api.get(f"/sites/{site_id}/stats/clients") or []
    return len(clients) if isinstance(clients, list) else 0

def get_all_sites_sle(api: MistAPI, org_id: str, hours: int = 24) -> dict:
    """
    Fetch Successful Connect SLE for ALL sites at once using the org-level
    insights endpoint.  Handles pagination automatically.

    Endpoint: GET /orgs/{org_id}/insights/sites-sle?sle=wifi&start=...&end=...
    Returns: dict  {site_id -> sle_value_float}
    """
    end_ts   = int(time.time())
    start_ts = end_ts - (hours * 3600)
    result   = {}

    page     = 1
    per_page = 100
    while True:
        params = {
            "sle":   "wifi",
            "start": start_ts,
            "end":   end_ts,
            "limit": per_page,
            "page":  page,
        }
        data = api.get(f"/orgs/{org_id}/insights/sites-sle", params=params)
        if data is None:
            break

        # Response may be a list or a dict containing a list
        rows = data if isinstance(data, list) else data.get("results", [])
        if not rows:
            break

        for row in rows:
            if not isinstance(row, dict):
                continue
            sid = row.get("site_id") or row.get("id")
            if not sid:
                continue
            # The SLE value is nested in a 'sle' dict keyed by metric name,
            # or may be a top-level 'value'/'successful-connect' key.
            sle_block = row.get("sle") or row.get("wifi") or {}
            val = None
            if isinstance(sle_block, dict):
                # Try common key variants
                for k in ("wifi-successful-connect", "successful-connect",
                          "successful_connect", "Successful Connect", "value"):
                    if k in sle_block:
                        val = sle_block[k]
                        break
                if val is None:
                    # Take the first numeric value we find
                    for v in sle_block.values():
                        if isinstance(v, (int, float)):
                            val = v
                            break
            # Also check top-level keys
            if val is None:
                for k in ("value", "successful-connect", "successful_connect"):
                    if k in row and isinstance(row[k], (int, float)):
                        val = row[k]
                        break
            if val is not None:
                raw = float(val)
                # Mist returns SLE as a fraction 0.0-1.0 where 1.0 = 100%.
                # Normalise to 0-100 percentage scale.
                pct = raw * 100.0 if raw <= 1.0 else raw
                result[sid] = pct
                debug(f"  SLE site {sid} raw={raw} pct={pct:.1f}%")

        if len(rows) < per_page:
            break
        page += 1

    log(f"  SLE data retrieved for {len(result)} site(s).", "OK")
    return result


# ---------------------------------------------------------------------------
# Best Practices definitions
# ---------------------------------------------------------------------------
BEST_PRACTICES = [
    {
        "id":   "arp_filter",
        "name": "ARP Filtering",
        "field": "arp_filter",
        "desired": True,
        "impact": (
            "ARP broadcast storms can degrade Wi-Fi performance in dense environments. "
            "Enabling ARP filtering causes the AP to proxy-ARP on behalf of known clients, "
            "dramatically reducing broadcast traffic over the air. Especially impactful in "
            "large-scale and IoT-heavy deployments."
        ),
        "remediation_key": "arp_filter",
    },
    {
        "id":   "limit_bcast",
        "name": "Multicast/Broadcast Filtering",
        "field": "limit_bcast",
        "desired": True,
        "impact": (
            "Broadcast and multicast frames must be received by every client, consuming "
            "valuable airtime. Enabling this filter suppresses unnecessary layer-2 floods, "
            "reducing channel utilisation and improving throughput and latency for all "
            "clients on that SSID."
        ),
        "remediation_key": "limit_bcast",
    },
    {
        "id":   "allow_ipv6_ndp",
        "name": "IPv6 NDP Disabled",
        "field": "allow_ipv6_ndp",
        "desired": False,
        "impact": (
            "IPv6 Neighbor Discovery Protocol (NDP) multicast traffic can saturate the "
            "wireless medium in dense environments. Unless the WLAN is explicitly designed "
            "for IPv6-only or dual-stack clients, disabling NDP reduces unnecessary "
            "multicast overhead and improves airtime efficiency for all connected clients."
        ),
        "remediation_key": "allow_ipv6_ndp",
    },
    {
        "id":   "dot11r",
        "name": "802.11r Fast Transition (Enterprise)",
        "field": None,
        "desired": True,
        "impact": (
            "802.11r Fast BSS Transition pre-authenticates clients with neighboring APs "
            "before they roam, reducing roam latency from >100 ms to <50 ms. Critical for "
            "voice, video, and real-time applications. Mist supports Hybrid 802.11r so "
            "legacy clients that do not support FT are unaffected. Applicable to "
            "WPA2/WPA3 Enterprise SSIDs only."
        ),
        "remediation_key": None,  # nested in auth object
    },
    {
        "id":   "duplicate_ssid",
        "name": "No Duplicate SSID Names",
        "field": None,
        "desired": True,
        "impact": (
            "Duplicate SSID names within the org can cause unexpected client behaviour, "
            "authentication loops, and makes troubleshooting harder. Each SSID should "
            "have a unique name unless there is an intentional design reason."
        ),
        "remediation_key": None,
    },
]


# ---------------------------------------------------------------------------
# Best-practice evaluation
# ---------------------------------------------------------------------------

def check_dot11r(wlan: dict):
    auth      = wlan.get("auth", {})
    auth_type = auth.get("type", "open").lower()
    if auth_type not in ENTERPRISE_AUTH_TYPES:
        return None  # N/A for non-enterprise
    disable_ft = auth.get("disable_ft", True)
    return not disable_ft  # disable_ft=False -> FT enabled -> compliant

def evaluate_wlan(wlan: dict) -> dict:
    """Return {bp_id: {'compliant': bool|None, 'current': any}} for one WLAN."""
    results = {}
    for bp in BEST_PRACTICES:
        bid = bp["id"]
        if bid == "duplicate_ssid":
            continue
        if bid == "dot11r":
            c   = check_dot11r(wlan)
            cur = wlan.get("auth", {}).get("disable_ft", "N/A")
        else:
            field   = bp["field"]
            cur     = wlan.get(field)
            desired = bp["desired"]
            if cur is not None:
                c = (cur == desired)
            elif desired is False:
                # Field absent = Mist default (not explicitly enabled) = compliant
                # when the best practice is to have the feature disabled.
                c = True
            else:
                # Field absent and desired=True → not enabled → non-compliant
                c = False
        results[bid] = {"compliant": c, "current": cur}
    return results

def is_non_compliant(bp_results: dict) -> bool:
    return any(v.get("compliant") is False for v in bp_results.values())


# ---------------------------------------------------------------------------
# Display helpers
# ---------------------------------------------------------------------------
PASS_STR = "\033[92m[PASS]\033[0m"
FAIL_STR = "\033[91m[FAIL]\033[0m"
NA_STR   = "\033[93m[ N/A]\033[0m"

def status_str(val):
    if val is True:  return PASS_STR
    if val is False: return FAIL_STR
    return NA_STR

def print_wlan_bp_table(wlan_name: str, bp_results: dict, show_passing=True):
    print(f"\n  WLAN: \033[1m{wlan_name}\033[0m")
    print(f"  {'Best Practice':<45} {'Status':<10} {'Current Value'}")
    print(f"  {'-'*45} {'-'*10} {'-'*30}")
    for bp in BEST_PRACTICES:
        bid = bp["id"]
        if bid == "duplicate_ssid":
            continue
        r    = bp_results.get(bid, {})
        comp = r.get("compliant")
        if not show_passing and comp is not False:
            continue
        cur  = str(r.get("current", ""))[:40]
        print(f"  {bp['name']:<45} {status_str(comp):<18} {cur}")

def print_best_practices_guide():
    section("WLAN Best Practices Reference Guide")
    for i, bp in enumerate(BEST_PRACTICES, 1):
        print(f"\n  {i}. \033[1m{bp['name']}\033[0m")
        words = bp["impact"].split()
        line  = "     "
        for w in words:
            if len(line) + len(w) + 1 > 72:
                print(line)
                line = "     " + w + " "
            else:
                line += w + " "
        if line.strip():
            print(line)


# ---------------------------------------------------------------------------
# Core data collection
# ---------------------------------------------------------------------------
def collect_all(api: MistAPI, org_id: str) -> dict:
    data = {
        "org_id":              org_id,
        "collected_at":        _ts(),
        "sites":               [],
        "org_wlans":           [],   # all enabled org-level WLANs
        "clients_before":      {},   # site_id -> count (collected at start)
        "clients_after":       None, # site_id -> count (collected after changes)
        "clients_total_before": 0,
        "clients_total_after":  0,
        "sle_before":          {},   # site_id -> SLE % (0-100)
    }

    # Sites
    data["sites"] = get_sites(api, org_id)

    # Org WLANs (all, then filter to enabled only)
    all_org_wlans = get_org_wlans(api, org_id)
    enabled_wlans = [w for w in all_org_wlans if w.get("enabled", True) is not False]
    disabled_count = len(all_org_wlans) - len(enabled_wlans)
    data["org_wlans"] = enabled_wlans
    if disabled_count:
        log(f"  Skipped {disabled_count} disabled WLAN(s).", "WARN")
    log(f"  Auditing {len(enabled_wlans)} enabled org WLAN(s).", "OK")

    # Client counts per site
    section("Loading Site Client Counts")
    sites = data["sites"]
    for i, site in enumerate(sites):
        sid = site["id"]
        progress(i + 1, len(sites), site.get("name", sid))
        data["clients_before"][sid] = get_site_client_count(api, sid)
    data["clients_total_before"] = sum(data["clients_before"].values())
    log(f"  Total wireless clients: {data['clients_total_before']}", "OK")

    # SLE for all sites (single paginated call)
    section("Loading Successful Connect SLE (all sites)")
    data["sle_before"] = get_all_sites_sle(api, org_id, hours=24)

    return data


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------
def report_sites(data: dict):
    section("All Sites")
    print(f"\n  {'#':<5} {'Site Name':<45} {'Site ID'}")
    print(f"  {'-'*5} {'-'*45} {'-'*36}")
    for i, site in enumerate(data["sites"], 1):
        print(f"  {i:<5} {site.get('name', '(unnamed)'):<45} {site['id']}")
    log(f"  Total sites: {len(data['sites'])}", "OK")

def report_org_wlans(data: dict):
    section("Org-Level Enabled WLANs")
    wlans = data["org_wlans"]
    if not wlans:
        log("  No enabled org WLANs found.", "WARN")
        return
    print(f"\n  {'#':<5} {'SSID':<35} {'Auth Type':<20} {'WLAN ID'}")
    print(f"  {'-'*5} {'-'*35} {'-'*20} {'-'*36}")
    for i, w in enumerate(wlans, 1):
        auth = w.get("auth", {}).get("type", "open")
        print(f"  {i:<5} {w.get('ssid','?'):<35} {auth:<20} {w.get('id','')}")
    log(f"  Total enabled org WLANs: {len(wlans)}", "OK")

def report_client_summary(data: dict, label: str = "Current"):
    """
    Print per-site client counts.
    label = "Before" when called at start, "After" when called post-changes.
    Uses data["clients_before"] always; call update_clients_after() first to
    populate data["clients_after"] and pass label="After".
    """
    section(f"Wireless Client Summary ({label})")
    counts = data.get("clients_after") if label == "After" else data["clients_before"]
    if counts is None:
        counts = data["clients_before"]
    total  = sum(counts.values()) if counts else 0
    label_key = "clients_total_after" if label == "After" else "clients_total_before"
    data[label_key] = total

    bar = "#" * 50
    print(f"\n  +{bar}+")
    print(f"  |  Total WLAN Clients ({label}): {total:<35}|")
    print(f"  +{bar}+\n")
    print(f"  {'Site':<45} {'Clients':>8}")
    print(f"  {'-'*45} {'-'*8}")
    for site in data["sites"]:
        cnt = counts.get(site["id"], 0)
        print(f"  {site.get('name', site['id']):<45} {cnt:>8}")


def collect_clients_after(api: MistAPI, data: dict):
    """Re-fetch per-site client counts and store in data['clients_after']."""
    data["clients_after"] = {}
    for site in data["sites"]:
        sid = site["id"]
        data["clients_after"][sid] = get_site_client_count(api, sid)
    data["clients_total_after"] = sum(data["clients_after"].values())

def report_sle_summary(data: dict):
    section("Successful Connect SLE - Last 24 Hours (All Sites)")
    sle_map  = data["sle_before"]
    below    = []
    no_data  = []
    print(f"\n  {'Site':<45} {'SLE (24h)':>10}  {'Status'}")
    print(f"  {'-'*45} {'-'*10}  {'-'*10}")
    for site in data["sites"]:
        sid  = site["id"]
        val  = sle_map.get(sid)
        sname = site.get("name", sid)
        if isinstance(val, float):
            display = f"{val:.1f}%"
            if val < SLE_THRESHOLD_PCT:
                status = "\033[91m[BELOW {:.0f}%]\033[0m".format(SLE_THRESHOLD_PCT)
                below.append((sname, val))
            else:
                status = "\033[92m[OK]\033[0m"
        else:
            display = "N/A"
            status  = "\033[93m[NO DATA]\033[0m"
            no_data.append(sname)
        print(f"  {sname:<45} {display:>10}  {status}")
    # Summary line
    print()
    if below:
        log(f"  {len(below)} site(s) below SLE threshold ({SLE_THRESHOLD_PCT:.0f}%):", "WARN")
        for sname, val in below:
            log(f"    - {sname}: {val:.1f}%", "WARN")
    else:
        log(f"  All sites meet the {SLE_THRESHOLD_PCT:.0f}% Successful Connect SLE threshold.", "OK")
    if no_data:
        log(f"  {len(no_data)} site(s) returned no SLE data (SLE may not be licensed/configured).", "WARN")

def report_bp_status(data: dict) -> list:
    """
    Evaluate all enabled org WLANs against best practices.
    Returns list of result dicts, one per WLAN.
    Also prints a summary table.
    """
    section("Best Practices Check - All Enabled Org WLANs")
    all_results = []
    non_compliant = []

    wlans = data["org_wlans"]
    if not wlans:
        log("  No WLANs to check.", "WARN")
        return all_results

    # Check for duplicate SSIDs across all org WLANs
    ssid_counts = defaultdict(int)
    for w in wlans:
        ssid_counts[w.get("ssid", "")] += 1
    dup_ssids = {s for s, c in ssid_counts.items() if c > 1}

    for wlan in wlans:
        bp_res = evaluate_wlan(wlan)
        ssid   = wlan.get("ssid", "?")
        # Add duplicate SSID check
        bp_res["duplicate_ssid"] = {
            "compliant": ssid not in dup_ssids,
            "current":   f"appears {ssid_counts[ssid]}x" if ssid in dup_ssids else "unique",
        }
        entry = {
            "wlan_id":    wlan.get("id", ""),
            "ssid":       ssid,
            "bp_results": bp_res,
            "wlan_obj":   wlan,
        }
        all_results.append(entry)
        if is_non_compliant(bp_res):
            non_compliant.append(entry)

    # Summary table
    print(f"\n  {'SSID':<35} {'ARP':^6} {'BCAST':^6} {'NDP':^6} "
          f"{'FT':^6} {'DupSS':^6}")
    print(f"  {'-'*35} {'-'*6} {'-'*6} {'-'*6} {'-'*6} {'-'*6}")
    bp_cols = ["arp_filter", "limit_bcast", "allow_ipv6_ndp",
               "dot11r", "duplicate_ssid"]
    for entry in all_results:
        row = f"  {entry['ssid']:<35}"
        for bid in bp_cols:
            comp = entry["bp_results"].get(bid, {}).get("compliant")
            sym  = "OK" if comp is True else ("FAIL" if comp is False else "N/A")
            color = "\033[92m" if comp is True else ("\033[91m" if comp is False else "\033[93m")
            row += f" {color}{sym:^6}\033[0m"
        print(row)

    log(f"\n  Checked {len(all_results)} WLAN(s): "
        f"{len(all_results)-len(non_compliant)} compliant, "
        f"{len(non_compliant)} non-compliant.", "OK")

    if non_compliant:
        section("WLANs NOT in Compliance with Best Practices")
        for entry in non_compliant:
            print_wlan_bp_table(entry["ssid"], entry["bp_results"], show_passing=False)
    else:
        log("  All WLANs are compliant with best practices.", "OK")

    return all_results


# ---------------------------------------------------------------------------
# Remediation
# ---------------------------------------------------------------------------
def apply_best_practices(api: MistAPI, org_id: str, bp_results: list) -> list:
    """
    Interactive per-WLAN, per-best-practice remediation.
    For each non-compliant WLAN, lists each failing BP and asks Update Y/N.
    Returns list of change records.
    """
    section("Update WLANs - Apply Best Practices")
    changes = []

    non_compliant = [r for r in bp_results if is_non_compliant(r["bp_results"])]
    if not non_compliant:
        log("  All WLANs are compliant - nothing to update.", "OK")
        return changes

    log(f"  {len(non_compliant)} non-compliant WLAN(s) found.", "WARN")

    for entry in non_compliant:
        ssid    = entry["ssid"]
        wlan_id = entry["wlan_id"]
        wlan    = entry["wlan_obj"]

        # Collect failing BPs
        failing = [(bp, entry["bp_results"][bp["id"]])
                   for bp in BEST_PRACTICES
                   if entry["bp_results"].get(bp["id"], {}).get("compliant") is False]

        if not failing:
            continue

        print(f"\n  {'='*65}")
        print(f"  WLAN: \033[1m{ssid}\033[0m  (id: {wlan_id})")
        print(f"  {len(failing)} best practice(s) not in compliance:\n")

        updated_payload = dict(wlan)  # accumulate changes
        applied_any     = False

        for bp, res in failing:
            bid     = bp["id"]
            cur_val = str(res.get("current", ""))

            print(f"    \033[91m[FAIL]\033[0m  {bp['name']}")
            print(f"           Current value : {cur_val}")
            print(f"           Impact        : {bp['impact'][:120]}...")

            if bid == "duplicate_ssid":
                log(f"    NOTE: Duplicate SSID '{ssid}' must be resolved manually in the Mist UI.", "WARN")
                changes.append({"ssid": ssid, "bp": bp["name"],
                                 "status": "manual-action-required"})
                continue

            if not ask_yn(f"    Update '{bp['name']}' on WLAN '{ssid}'?"):
                log(f"    Skipped '{bp['name']}' on '{ssid}'", "WARN")
                changes.append({"ssid": ssid, "bp": bp["name"],
                                 "status": "skipped"})
                continue

            rem_key = bp.get("remediation_key")
            if rem_key:
                updated_payload[rem_key] = bp["desired"]
            elif bid == "dot11r":
                auth = dict(updated_payload.get("auth", {}))
                auth["disable_ft"] = False
                updated_payload["auth"] = auth
            applied_any = True

        if applied_any:
            log(f"  Patching WLAN '{ssid}' (id: {wlan_id}) ...")
            result = api.put(f"/orgs/{org_id}/wlans/{wlan_id}", updated_payload)
            if result:
                log(f"  [OK] WLAN '{ssid}' updated successfully.", "OK")
                for bp, _ in failing:
                    if bp.get("remediation_key") or bp["id"] == "dot11r":
                        changes.append({"ssid": ssid, "bp": bp["name"],
                                         "status": "applied"})
            else:
                log(f"  [FAIL] Failed to patch WLAN '{ssid}' - see debug log.", "ERROR")
                for bp, _ in failing:
                    if bp.get("remediation_key") or bp["id"] == "dot11r":
                        changes.append({"ssid": ssid, "bp": bp["name"],
                                         "status": "failed"})

    return changes


# ---------------------------------------------------------------------------
# Duplicate SSID report
# ---------------------------------------------------------------------------
def report_duplicate_ssids(data: dict):
    section("Duplicate SSID Names - Org WLANs")
    ssid_counts = defaultdict(int)
    for w in data["org_wlans"]:
        ssid_counts[w.get("ssid", "")] += 1
    dupes = {s: c for s, c in ssid_counts.items() if c > 1}
    if dupes:
        for ssid, cnt in dupes.items():
            log(f"  Duplicate SSID: '{ssid}' appears {cnt}x", "WARN")
    else:
        log("  No duplicate SSID names found.", "OK")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def export_excel(data: dict, bp_results: list, output_dir: str):
    if not XLSX_AVAILABLE:
        log("Excel export skipped - openpyxl not installed.", "WARN")
        return None

    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"mist_bp_report_{ts}.xlsx")
    wb   = openpyxl.Workbook()

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(color="FFFFFF", bold=True)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    na_fill   = PatternFill("solid", fgColor="FFEB9C")
    thin      = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    def sh(cell):
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    def sd(cell, fill=None):
        if fill: cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True)

    bp_ids   = [b["id"] for b in BEST_PRACTICES]
    bp_names = [b["name"] for b in BEST_PRACTICES]

    # Tab 1: Org WLANs - Best Practices
    ws1 = wb.active
    ws1.title = "Org WLAN Best Practices"
    headers = ["SSID", "Auth Type", "WLAN ID"] + bp_names
    for col, h in enumerate(headers, 1):
        sh(ws1.cell(row=1, column=col, value=h))
    row = 2
    for r in bp_results:
        ws1.cell(row=row, column=1, value=r["ssid"])
        ws1.cell(row=row, column=2, value=r["wlan_obj"].get("auth", {}).get("type", "open"))
        ws1.cell(row=row, column=3, value=r["wlan_id"])
        for ci, bid in enumerate(bp_ids, 4):
            comp  = r["bp_results"].get(bid, {}).get("compliant")
            val   = "PASS" if comp is True else ("FAIL" if comp is False else "N/A")
            fill  = pass_fill if comp is True else (fail_fill if comp is False else na_fill)
            sd(ws1.cell(row=row, column=ci, value=val), fill)
        row += 1
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 22
    ws1.freeze_panes = "A2"

    # Tab 2: Non-Compliant WLANs
    ws2 = wb.create_sheet("Non-Compliant WLANs")
    nc_headers = ["SSID", "Best Practice", "Current Value", "Status"]
    for col, h in enumerate(nc_headers, 1):
        sh(ws2.cell(row=1, column=col, value=h))
    row2 = 2
    for r in bp_results:
        for bp in BEST_PRACTICES:
            bid  = bp["id"]
            comp = r["bp_results"].get(bid, {}).get("compliant")
            if comp is False:
                cur = str(r["bp_results"].get(bid, {}).get("current", ""))
                ws2.cell(row=row2, column=1, value=r["ssid"])
                ws2.cell(row=row2, column=2, value=bp["name"])
                ws2.cell(row=row2, column=3, value=cur)
                sd(ws2.cell(row=row2, column=4, value="FAIL"), fail_fill)
                row2 += 1
    for col in range(1, len(nc_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 30
    ws2.freeze_panes = "A2"

    # Tab 3: Site SLE Summary
    ws3 = wb.create_sheet("Site SLE Summary")
    sle_headers = ["Site Name", "Site ID", "Clients (current)",
                   "SLE Successful Connect (24h %)",
                   f"Below {SLE_THRESHOLD_PCT:.0f}% Threshold?"]
    for col, h in enumerate(sle_headers, 1):
        sh(ws3.cell(row=1, column=col, value=h))
    row3 = 2
    for site in data["sites"]:
        sid = site["id"]
        val = data["sle_before"].get(sid)
        display = f"{val:.1f}%" if isinstance(val, float) else "N/A"
        below_flag = "YES" if isinstance(val, float) and val < SLE_THRESHOLD_PCT else \
                     ("NO" if isinstance(val, float) else "N/A")
        ws3.cell(row=row3, column=1, value=site.get("name", sid))
        ws3.cell(row=row3, column=2, value=sid)
        ws3.cell(row=row3, column=3, value=data["clients_before"].get(sid, 0))
        ws3.cell(row=row3, column=4, value=display)
        c5 = ws3.cell(row=row3, column=5, value=below_flag)
        if below_flag == "YES":
            sd(c5, fail_fill)
        elif below_flag == "NO":
            sd(c5, pass_fill)
        row3 += 1
    for col in range(1, len(sle_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 30
    ws3.freeze_panes = "A2"

    wb.save(path)
    log(f"  Excel report saved -> {path}", "OK")
    return path


# ---------------------------------------------------------------------------
# Post-change verification
# ---------------------------------------------------------------------------
def post_change_verification(api: MistAPI, org_id: str, data: dict):
    section("Post-Change Verification")
    log("  Waiting 30 seconds for changes to propagate ...")
    time.sleep(30)

    log("  Re-checking client counts ...")
    collect_clients_after(api, data)
    for site in data["sites"]:
        sid    = site["id"]
        before = data["clients_before"].get(sid, 0)
        after  = data["clients_after"].get(sid, 0)
        delta  = after - before
        sym    = "^" if delta > 0 else ("v" if delta < 0 else "=")
        log(f"  {site.get('name', sid):<45} before={before}  after={after}  {sym}{abs(delta)}")
    log(f"  Org total: before={data['clients_total_before']}  after={data['clients_total_after']}")

    log("  Re-checking SLE for past 60 minutes ...")
    sle_after_map = get_all_sites_sle(api, org_id, hours=1)
    for site in data["sites"]:
        sid        = site["id"]
        val_before = data["sle_before"].get(sid)   # already 0-100 pct
        val_after  = sle_after_map.get(sid)         # already 0-100 pct
        b_str = f"{val_before:.1f}%" if isinstance(val_before, float) else "N/A"
        a_str = f"{val_after:.1f}%"  if isinstance(val_after,  float) else "N/A"
        flag  = " \033[91m[BELOW {:.0f}%]\033[0m".format(SLE_THRESHOLD_PCT) \
                if isinstance(val_after, float) and val_after < SLE_THRESHOLD_PCT else ""
        log(f"  {site.get('name', sid):<45}  SLE 24h={b_str}  SLE 1h={a_str}{flag}")


# ---------------------------------------------------------------------------
# Midnight automation
# ---------------------------------------------------------------------------
def schedule_midnight_run():
    section("Midnight Automation Setup")
    script_path = os.path.abspath(__file__)
    log("  To run this script automatically every midnight:")
    log(f"  Linux/macOS cron:  0 0 * * *  python3 \"{script_path}\" --auto")
    log(f"  Windows Task Scheduler:  python \"{script_path}\" --auto")
    log("  For manual interactive runs use:  python3 script.py --main  (or no flag)")
    log("  Required environment variables for --auto mode:")
    log("    MIST_CLOUD   = 1-6 (cloud number)")
    log("    MIST_ORG_ID  = <your org UUID>")
    log("    MIST_TOKEN   = <SuperUser API token>")


# ---------------------------------------------------------------------------
# Auto mode
# ---------------------------------------------------------------------------
def run_auto_mode(base_url: str, org_id: str, token: str):
    api = MistAPI(base_url, token)
    verify_auth(api)

    data       = collect_all(api, org_id)
    bp_results = report_bp_status(data)
    changes    = apply_best_practices(api, org_id, bp_results)
    log(f"  Auto-mode: {len(changes)} change(s) applied.", "OK")

    post_change_verification(api, org_id, data)   # populates clients_after
    output_dir = os.path.dirname(os.path.abspath(__file__))
    export_excel(data, bp_results, output_dir)
    delta = data["clients_total_after"] - data["clients_total_before"]
    log(f"  Clients before : {data['clients_total_before']}")
    log(f"  Clients after  : {data['clients_total_after']} ({'+' if delta>=0 else ''}{delta})")
    log(f"  API calls: {api.call_count}")
    log(f"  Elapsed  : {_elapsed()}")
    log(f"  Log      : {LOG_FILE}")
    log(f"  Debug    : {DEBUG_LOG}")


# ---------------------------------------------------------------------------
# Main interactive flow
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Mist WLAN Best Practices Automation v2")
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument("--main", action="store_true",
                        help="Interactive mode - prompt for credentials and walk through all checks (default)")
    mode_group.add_argument("--auto", action="store_true",
                        help="Non-interactive mode (uses MIST_CLOUD, MIST_ORG_ID, MIST_TOKEN env vars)")
    args = parser.parse_args()

    section(f"Juniper Mist WLAN Best Practices Automation  v{SCRIPT_VERSION}")
    log(f"  Start time : {_ts()}")
    log(f"  Log file   : {LOG_FILE}")
    log(f"  Debug log  : {DEBUG_LOG}")

    # -- Auto mode -------------------------------------------------------------
    if args.auto:
        cloud_key = os.environ.get("MIST_CLOUD", "1")
        cloud_obj = MIST_CLOUDS.get(cloud_key, MIST_CLOUDS["1"])
        org_id    = os.environ.get("MIST_ORG_ID", "")
        token     = os.environ.get("MIST_TOKEN", "")
        if not org_id or not token:
            sys.exit("ERROR: MIST_ORG_ID and MIST_TOKEN env vars are required in --auto mode.")
        run_auto_mode(cloud_obj["base"], org_id, token)
        return

    # -- Interactive mode (--main or no flag) ----------------------------------
    # args.main is True when --main is explicitly passed; if neither flag is
    # given the script defaults to interactive mode as well.

    # 1. Credentials - SuperUser token only
    base_url = prompt_cloud()
    org_id   = prompt_org_id()
    token    = prompt_token("SuperUser API Token")

    api = MistAPI(base_url, token)
    verify_auth(api)

    # 2. Collect all data
    data = collect_all(api, org_id)

    # 3. List all sites
    report_sites(data)

    # 4. List all enabled org WLANs
    report_org_wlans(data)

    # 5. Client summary (Before)
    report_client_summary(data, label="Before")

    # 6. SLE summary (all sites)
    report_sle_summary(data)

    # 7. Best practices guide (optional)
    if ask_yn("Do you want to view the WLAN Best Practices reference guide?"):
        print_best_practices_guide()

    # 8. Check best practices + list non-compliant WLANs
    bp_results = report_bp_status(data)

    # 9. Offer to update non-compliant WLANs
    non_compliant = [r for r in bp_results if is_non_compliant(r["bp_results"])]
    changes = []
    if non_compliant:
        if ask_yn(f"Do you want to update the {len(non_compliant)} non-compliant WLAN(s) now?"):
            changes = apply_best_practices(api, org_id, bp_results)
            log(f"  {len(changes)} change record(s) logged.", "OK")
            post_change_verification(api, org_id, data)
    else:
        log("  All WLANs are compliant - no updates needed.", "OK")

    # 10. Collect clients-after (if not already done by post_change_verification)
    if data["clients_after"] is None:
        log("  Collecting final client counts ...")
        collect_clients_after(api, data)
    report_client_summary(data, label="After")

    # 11. Midnight automation
    if ask_yn("Do you want to set up automatic nightly best-practices checking?"):
        schedule_midnight_run()

    # 12. Excel export
    output_dir  = os.path.dirname(os.path.abspath(__file__))
    excel_path  = export_excel(data, bp_results, output_dir)

    # 13. Final summary
    section("Run Summary")
    client_delta = data["clients_total_after"] - data["clients_total_before"]
    delta_str    = f"({'+' if client_delta >= 0 else ''}{client_delta})"
    log(f"  Sites                          : {len(data['sites'])}")
    log(f"  Org WLANs (enabled)            : {len(data['org_wlans'])}")
    log(f"  Total WLAN Clients (Before)    : {data['clients_total_before']}")
    log(f"  Total WLAN Clients (After)     : {data['clients_total_after']} {delta_str}")
    log(f"  Non-compliant WLANs            : {len(non_compliant)}")
    log(f"  Changes applied                : {len(changes)}")
    log(f"  API calls this session         : {api.call_count}")
    log(f"  Total elapsed time             : {_elapsed()}")
    log(f"  Main log file                  : {LOG_FILE}")
    log(f"  Debug log file                 : {DEBUG_LOG}")
    if excel_path:
        log(f"  Excel report          : {excel_path}")


if __name__ == "__main__":
    main()
